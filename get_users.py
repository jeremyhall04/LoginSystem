from getpass import getpass
import os
from openpyxl import load_workbook, Workbook
from tkinter import messagebox

valid_user = False


def check_user(user, valid_user):
    wb = load_workbook("database.xlsx")
    ws = wb.active
    # try:
    for row in ws.iter_rows(
        min_row=0, max_row=200, min_col=0, max_col=2, values_only=True
    ):
        if row[0] == user:
            pw = row[1]
            valid_user = True
    if valid_user != True:
        pw = None
        print("     Username not in database. Please try again.")
        messagebox.showinfo(
            "Incorrect Username", "The username you entered is not in our database"
        )

    return valid_user, pw


def create_user():
    username = input("Enter a username: ")
    password = input("Enter a passowrd: ")
    inc_pass = True
    while inc_pass == True:
        re_password = input("Please re-enter your password: ")
        if password != re_password:
            print("     The passwords you entered did not match. Please try again")
        else:
            inc_pass = False

    wb = load_workbook("database.xlsx")
    ws = wb.active
    ws.append([username, password])
    wb.save("database.xlsx")

    print("\nAcount successfully created!\n")
    return username


def create_user(username, password):
    wb = load_workbook("database.xlsx")
    ws = wb.active
    ws.append([username, password])
    wb.save("database.xlsx")
    print("\nAcount successfully created!\n")
    return username
